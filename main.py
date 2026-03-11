from access_sheets import read_sheet,display_all_columns #access sheets will return the data stored inside each worksheet
from find_highest_payment_mode_total import read_excel #this read_excel function will read the excel file and returns all the sheet names stores in the workbook
import pandas as pd
import seaborn as sns
import os
import matplotlib.pyplot as plt
from access_sheets import work_sheets,all_sheets


display_all_columns() # displaying all columns of a dataframe

class FinancialAnalysis: # this class will display all records of the dataset i have in textual, graphical manner
    
    def __init__(self):
        self.excel_file=read_excel() #read the workbook
        self.access_sheet_data(work_sheets[-2])
        print(f'accessing data of sheet {self.sheet[1]}\n')
        # print(self.selected_df)
        return None
    # access sheet data by name and returning only relevant columns
    def access_sheet_data(self,sheet_name:str)->object:
        self.sheet=read_sheet(sheet_name) #returns the worksheet data and its name
        self.df=pd.DataFrame(self.sheet[0]) # converting sheets into dataframe
        self.selected_df=self.df[['sr.no.', 'Amount', 'Category', 'Description', 'Place', 'Payment-Mode', 'Date']] # choosing only those columns which are highly relevant for further work
        return self.selected_df
    # calculating expenses according to category (which category eats most of my budget)
    def show_category_expenses(self,sheets: str)-> object: #generate a graph showing which category consumed the most of my budget over the period
        self.df=self.access_sheet_data(sheet_name=sheets)
        self.category_table=self.df.groupby('Category')['Amount'].sum() #category wise sum of expenses that which category has consumed most of the budget
        return self.category_table 
    # plot graph based on categories and their amount sum
    def plot_graph(self)->None:
        self.category_table.plot(kind="bar")
        plt.xlabel('Category')
        plt.ylabel('Expenses')
        plt.title('Expenses by Category')
        plt.show()
        return None
    # function returns the difference in amount city wise
    def spendings_difference_city_wise(self)-> object:
        self.spendings_difference=self.selected_df.groupby('Place')['Amount'].sum()
        self.spendings_difference.plot(kind="bar")
        plt.xlabel('Place')
        plt.ylabel('Amount')
        plt.title(f'Expenses by City for the month of {self.sheet[1].removesuffix("-spendings")}')
        plt.show()
        return None
    # check which months has the highest and which month has the lowest spending amount
    def highest_lowest_spending_month(self):
        from openpyxl import load_workbook
        wb = load_workbook(r'C:\Users\HP\Desktop\finance_project\spendings-record.xlsx')
        ws = wb[all_sheets[-2]]
        table=ws.tables["past6monthsexpense"] #accessing table which is relevant for calculating past 6 months expense
        data = ws[table.ref]
        df= pd.DataFrame([[cell.value for cell in row] for row in data])
        return df
    
    # find how much amount of money is spent under certain time period (5 or 10 or 20 days etc)
    def spending_amount_time_period(self,start_date:str,end_date:str):
        _addition=0
        self.date=[]
        # print(_addition)
        
        for date in self.selected_df['Date']: #returns all the date period for a specific sheet
            self.date.append(str(date).removesuffix('00:00:00')) #remove unncecessary string from end       
        total = self.selected_df[(self.selected_df["Date"] >= start_date) & (self.selected_df["Date"] <= end_date)]["Amount"].sum()
        return total
    

f1=FinancialAnalysis()
start = "2026-01-13"
end = "2026-02-09"
amount=f1.spending_amount_time_period("2026-01-13","2026-02-09")

print(f'the amount spent between {start} and {end} is rupees {amount}')
# print(f1.spendings_difference_city_wise())
print(f1.show_category_expenses(all_sheets[2]))
print(f1.plot_graph())



